import openpyxl
from product.models import Product
from product.serializers import ProductSerializer
import json
import os
from django.conf import settings
from datetime import datetime



def convert_expiry_to_upper_bound(expiry_range):
    # Check if the value is already an integer or a float
    if isinstance(expiry_range, (int, float)):
        return expiry_range
    # Check if the value contains '-'
    if isinstance(expiry_range, str) and '-' in expiry_range:
        try:
            # Attempt to split the string by hyphen
            upper_bound = int(str(expiry_range).split('-')[1])
            return upper_bound
        except (IndexError, ValueError, AttributeError):
            return None
    return None
        

def populate_db():
    # populate product model from excel file
    # open the excel file
    excel_path = os.path.join(settings.BASE_DIR, 'product', 'product_list.xlsx')
        
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return
    wb = openpyxl.load_workbook(excel_path)
    # select the sheet
    sheet = wb['Sheet1']
    # get the max row count
    max_row = sheet.max_row
    # iterate over all the rows
    for i in range(2, max_row + 1):
        # get the data from each cell
        print(f'Updating: {sheet.cell(row=i, column=1).value}')
        product_name = sheet.cell(row=i, column=1).value
        product_category = sheet.cell(row=i, column=2).value
        product_tags = parse_tags(sheet.cell(row=i, column=3).value)
        allergy_tags = parse_allergy_tags(sheet.cell(row=i, column=4).value)
        product_standard_expiry_days = convert_expiry_to_upper_bound(sheet.cell(row=i, column=5).value)
        # create or update a product object
        product = Product.objects.filter(name=product_name).first()
        if product:
            product_serializer = ProductSerializer(product, data={
                'name': product_name,
                'category': product_category,
                'tags': product_tags,
                'standard_expiry_days': product_standard_expiry_days,
                'allergy_tags': allergy_tags,
            })
        else:
            product_serializer = ProductSerializer(data={
                'name': product_name,
                'category': product_category,
                'tags': product_tags,
                'standard_expiry_days': product_standard_expiry_days,
                'allergy_tags': allergy_tags,
            })
        # check if the data is valid
        if product_serializer.is_valid():
            # save the data
            product_serializer.save()
        else:
            print(product_serializer.errors)

def parse_tags(tags):
    if isinstance(tags, list):
        return tags
    else:
        tags = tags.replace('[', '').replace(']', '')
        if tags:
            return [tag.strip() for tag in tags.split(',')]
    return []

def parse_allergy_tags(allergy_tags):
    if isinstance(allergy_tags, list):
        return allergy_tags
    else:
        allergy_tags = allergy_tags.replace('[', '').replace(']', '')
        if allergy_tags:
            return [tag.strip() for tag in allergy_tags.split(' ')]
    return []